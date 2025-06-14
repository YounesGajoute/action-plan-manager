# ===================================================================
# backend/services/excel_service.py - Excel Import/Export Service
# ===================================================================

import os
import logging
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.exc import IntegrityError
from flask import current_app

from app.models import Task, User, SyncStatus
from app import db

logger = logging.getLogger(__name__)

class ExcelService:
    """Service for handling Excel import/export operations"""
    
    # Excel column mapping based on the actual file structure
    EXCEL_COLUMNS = {
        'Date': 'date_created',
        'PO': 'po_number',
        'Catégorie': 'category',
        'Action ': 'action_description',  # Note: trailing space in Excel
        'Colonne1': 'colonne1',
        'Customer': 'customer',
        'Requester': 'requester',
        'Techmac Resp': 'responsible',
        'Dead line ': 'deadline',  # Note: trailing space in Excel
        'Status': 'status',
        'Note': 'notes',
        'Installation/F': 'installation_flag',
        'Réparation': 'reparation_flag',
        'Développement': 'developpement_flag',
        'Livraison ': 'livraison_flag'  # Note: trailing space in Excel
    }
    
    STATUS_MAPPING = {
        'done': 'Terminé',
        'completed': 'Terminé',
        'finished': 'Terminé',
        'terminé': 'Terminé',
        'fini': 'Terminé',
        'complete': 'Terminé',
        'pending': 'En Attente',
        'waiting': 'En Attente',
        'en attente': 'En Attente',
        'attente': 'En Attente',
        'wait': 'En Attente',
        'in-progress': 'En Cours',
        'in progress': 'En Cours',
        'progress': 'En Cours',
        'en cours': 'En Cours',
        'cours': 'En Cours',
        'working': 'En Cours',
        'cancelled': 'Annulé',
        'canceled': 'Annulé',
        'annulé': 'Annulé',
        'annule': 'Annulé',
        'cancel': 'Annulé',
        'on-hold': 'En Pause',
        'on hold': 'En Pause',
        'hold': 'En Pause',
        'pause': 'En Pause',
        'en pause': 'En Pause',
        'paused': 'En Pause'
    }
    
    CATEGORY_MAPPING = {
        'installation': 'Installation',
        'install': 'Installation',
        'installing': 'Installation',
        'setup': 'Installation',
        'reparation': 'Réparation',
        'réparation': 'Réparation',
        'repair': 'Réparation',
        'fix': 'Réparation',
        'fixing': 'Réparation',
        'maintenance': 'Réparation',
        'developpement': 'Développement',
        'développement': 'Développement',
        'development': 'Développement',
        'dev': 'Développement',
        'programming': 'Développement',
        'coding': 'Développement',
        'livraison': 'Livraison',
        'delivery': 'Livraison',
        'shipping': 'Livraison',
        'transport': 'Livraison',
        'expedition': 'Livraison',
        'commercial': 'Commercial',
        'sales': 'Commercial',
        'vente': 'Commercial',
        'marketing': 'Commercial',
        'business': 'Commercial'
    }
    
    @staticmethod
    def validate_excel_file(file_path: str) -> Tuple[bool, str]:
        """Validate Excel file format and structure"""
        try:
            if not os.path.exists(file_path):
                return False, "File does not exist"
            
            # Check file extension
            if not file_path.lower().endswith(('.xlsx', '.xls')):
                return False, "Invalid file format. Please use .xlsx or .xls"
            
            # Try to load the workbook
            workbook = load_workbook(file_path, read_only=True)
            
            if not workbook.sheetnames:
                return False, "Excel file contains no sheets"
            
            # Get the first sheet
            sheet = workbook.active
            
            # Check if sheet has data
            if sheet.max_row < 2:
                return False, "Excel file contains no data rows"
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1] if cell.value]
            
            # Check for required columns
            required_columns = ['Action ', 'Customer', 'Requester', 'Techmac Resp']
            missing_columns = []
            
            for required_col in required_columns:
                if not any(required_col.lower() in str(header).lower() for header in headers):
                    missing_columns.append(required_col)
            
            if missing_columns:
                return False, f"Missing required columns: {', '.join(missing_columns)}"
            
            workbook.close()
            return True, "File is valid"
            
        except Exception as e:
            logger.error(f"Error validating Excel file: {str(e)}")
            return False, f"Error reading Excel file: {str(e)}"
    
    @staticmethod
    def parse_excel_date(date_value) -> Optional[datetime]:
        """Parse Excel date values into datetime objects"""
        if not date_value:
            return None
        
        try:
            if isinstance(date_value, datetime):
                return date_value
            elif isinstance(date_value, str):
                # Handle DD/MM/YY or DD/MM/YYYY format
                if '/' in date_value:
                    parts = date_value.strip().split('/')
                    if len(parts) == 3:
                        day, month, year = parts
                        # Convert 2-digit year to 4-digit
                        if len(year) == 2:
                            year_int = int(year)
                            if year_int <= 30:
                                year = '20' + year
                            else:
                                year = '19' + year
                        
                        return datetime.strptime(f"{day.zfill(2)}/{month.zfill(2)}/{year}", "%d/%m/%Y")
                else:
                    # Try ISO format
                    return datetime.fromisoformat(date_value)
            else:
                # Handle Excel serial date number
                from datetime import timedelta
                return datetime(1900, 1, 1) + timedelta(days=float(date_value) - 2)
                
        except Exception as e:
            logger.warning(f"Could not parse date '{date_value}': {str(e)}")
            return None
    
    @staticmethod
    def normalize_status(status: str) -> str:
        """Normalize status values from Excel"""
        if not status:
            return 'En Attente'
        
        status_lower = str(status).lower().strip()
        return ExcelService.STATUS_MAPPING.get(status_lower, status)
    
    @staticmethod
    def infer_category(row_data: Dict) -> Optional[str]:
        """Infer category from Excel data"""
        # Check category column first
        if row_data.get('Catégorie'):
            category_lower = str(row_data['Catégorie']).lower().strip()
            mapped_category = ExcelService.CATEGORY_MAPPING.get(category_lower)
            if mapped_category:
                return mapped_category
        
        # Check flag columns
        if row_data.get('Installation/F'):
            return 'Installation'
        elif row_data.get('Réparation'):
            return 'Réparation'
        elif row_data.get('Développement'):
            return 'Développement'
        elif row_data.get('Livraison '):
            return 'Livraison'
        
        # Check action description for keywords
        action = str(row_data.get('Action ', '')).lower()
        if 'commercial' in action:
            return 'Commercial'
        elif 'install' in action:
            return 'Installation'
        elif 'repair' in action or 'fix' in action:
            return 'Réparation'
        elif 'develop' in action or 'dev' in action:
            return 'Développement'
        elif 'deliver' in action or 'ship' in action:
            return 'Livraison'
        
        return None
    
    @staticmethod
    def is_flag_set(value) -> bool:
        """Check if Excel flag column is set"""
        if not value:
            return False
        
        str_value = str(value).lower().strip()
        return str_value in ['true', '1', 'yes', 'oui', 'x', '✓', 'checked'] or \
               (str_value != '' and str_value not in ['0', 'false', 'no', 'non'])
    
    @staticmethod
    def import_from_excel(file_path: str, user_id: str) -> Dict:
        """Import tasks from Excel file"""
        try:
            # Validate file first
            is_valid, validation_message = ExcelService.validate_excel_file(file_path)
            if not is_valid:
                return {
                    'success': False,
                    'error': validation_message,
                    'imported': 0,
                    'updated': 0,
                    'errors': []
                }
            
            # Load workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            
            # Track statistics
            imported_count = 0
            updated_count = 0
            errors = []
            processed_rows = 0
            
            # Process each row
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Create row data dictionary
                    row_data = dict(zip(headers, row))
                    
                    # Skip if no action description
                    if not row_data.get('Action '):
                        continue
                    
                    processed_rows += 1
                    
                    # Parse and clean data
                    task_data = ExcelService._parse_row_data(row_data)
                    
                    # Check if task exists (by PO number and action if PO exists)
                    existing_task = None
                    if task_data.get('po_number'):
                        existing_task = Task.query.filter_by(
                            po_number=task_data['po_number'],
                            action_description=task_data['action_description']
                        ).first()
                    
                    if existing_task:
                        # Update existing task
                        ExcelService._update_task_from_data(existing_task, task_data, user_id)
                        updated_count += 1
                        logger.info(f"Updated task {existing_task.id} from Excel row {row_num}")
                    else:
                        # Create new task
                        task = ExcelService._create_task_from_data(task_data, user_id)
                        db.session.add(task)
                        imported_count += 1
                        logger.info(f"Created new task from Excel row {row_num}")
                
                except Exception as e:
                    error_msg = f"Row {row_num}: {str(e)}"
                    errors.append(error_msg)
                    logger.error(f"Error processing Excel row {row_num}: {str(e)}")
                    continue
            
            # Commit all changes
            db.session.commit()
            
            # Create sync status record
            sync_status = SyncStatus(
                sync_type='excel_import',
                status='success',
                message=f'Excel import completed: {imported_count} imported, {updated_count} updated',
                items_processed=processed_rows,
                items_imported=imported_count,
                items_updated=updated_count,
                items_failed=len(errors),
                completed_at=datetime.utcnow()
            )
            db.session.add(sync_status)
            db.session.commit()
            
            workbook.close()
            
            logger.info(f"Excel import completed: {imported_count} imported, {updated_count} updated, {len(errors)} errors")
            
            return {
                'success': True,
                'imported': imported_count,
                'updated': updated_count,
                'errors': errors,
                'processed': processed_rows
            }
            
        except Exception as e:
            logger.error(f"Error importing Excel file: {str(e)}")
            db.session.rollback()
            return {
                'success': False,
                'error': str(e),
                'imported': 0,
                'updated': 0,
                'errors': []
            }
    
    @staticmethod
    def _parse_row_data(row_data: Dict) -> Dict:
        """Parse and clean row data from Excel"""
        task_data = {}
        
        # Parse date fields
        date_created = ExcelService.parse_excel_date(row_data.get('Date'))
        deadline = ExcelService.parse_excel_date(row_data.get('Dead line '))
        
        # Map Excel columns to task fields
        task_data.update({
            'po_number': str(row_data.get('PO', '')).strip() if row_data.get('PO') else None,
            'date_created': date_created or datetime.utcnow(),
            'category': ExcelService.infer_category(row_data),
            'action_description': str(row_data.get('Action ', '')).strip(),
            'colonne1': str(row_data.get('Colonne1', '')).strip() if row_data.get('Colonne1') else None,
            'customer': str(row_data.get('Customer', '')).strip(),
            'requester': str(row_data.get('Requester', '')).strip(),
            'responsible': str(row_data.get('Techmac Resp', '')).strip(),
            'deadline': deadline,
            'status': ExcelService.normalize_status(row_data.get('Status', 'En Attente')),
            'notes': str(row_data.get('Note', '')).strip() if row_data.get('Note') else None,
            'installation_flag': ExcelService.is_flag_set(row_data.get('Installation/F')),
            'reparation_flag': ExcelService.is_flag_set(row_data.get('Réparation')),
            'developpement_flag': ExcelService.is_flag_set(row_data.get('Développement')),
            'livraison_flag': ExcelService.is_flag_set(row_data.get('Livraison '))
        })
        
        return task_data
    
    @staticmethod
    def _create_task_from_data(task_data: Dict, user_id: str) -> Task:
        """Create a new Task from parsed data"""
        return Task(
            po_number=task_data.get('po_number'),
            date_created=task_data.get('date_created'),
            category=task_data.get('category'),
            action_description=task_data.get('action_description'),
            colonne1=task_data.get('colonne1'),
            customer=task_data.get('customer'),
            requester=task_data.get('requester'),
            responsible=task_data.get('responsible'),
            deadline=task_data.get('deadline'),
            status=task_data.get('status'),
            notes=task_data.get('notes'),
            installation_flag=task_data.get('installation_flag', False),
            reparation_flag=task_data.get('reparation_flag', False),
            developpement_flag=task_data.get('developpement_flag', False),
            livraison_flag=task_data.get('livraison_flag', False),
            created_by=user_id
        )
    
    @staticmethod
    def _update_task_from_data(task: Task, task_data: Dict, user_id: str):
        """Update existing task with new data"""
        # Update basic fields
        task.customer = task_data.get('customer', task.customer)
        task.requester = task_data.get('requester', task.requester)
        task.responsible = task_data.get('responsible', task.responsible)
        task.status = task_data.get('status', task.status)
        task.notes = task_data.get('notes', task.notes)
        task.deadline = task_data.get('deadline', task.deadline)
        task.category = task_data.get('category', task.category)
        task.colonne1 = task_data.get('colonne1', task.colonne1)
        
        # Update flags
        task.installation_flag = task_data.get('installation_flag', task.installation_flag)
        task.reparation_flag = task_data.get('reparation_flag', task.reparation_flag)
        task.developpement_flag = task_data.get('developpement_flag', task.developpement_flag)
        task.livraison_flag = task_data.get('livraison_flag', task.livraison_flag)
        
        # Update system fields
        task.updated_by = user_id
        task.updated_at = datetime.utcnow()
    
    @staticmethod
    def export_to_excel(tasks: List[Task], file_path: str) -> bool:
        """Export tasks to Excel file"""
        try:
            # Create workbook and worksheet
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Tasks"
            
            # Define headers (matching Excel structure)
            headers = [
                'Date', 'PO', 'Catégorie', 'Action', 'Colonne1', 'Customer',
                'Requester', 'Techmac Resp', 'Dead line', 'Status', 'Note',
                'Installation/F', 'Réparation', 'Développement', 'Livraison'
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_num, task in enumerate(tasks, 2):
                row_data = [
                    task.date_created.strftime('%d/%m/%Y') if task.date_created else '',
                    task.po_number or '',
                    task.category or '',
                    task.action_description or '',
                    task.colonne1 or '',
                    task.customer or '',
                    task.requester or '',
                    task.responsible or '',
                    task.deadline.strftime('%d/%m/%Y') if task.deadline else '',
                    task.status or '',
                    task.notes or '',
                    'X' if task.installation_flag else '',
                    'X' if task.reparation_flag else '',
                    'X' if task.developpement_flag else '',
                    'X' if task.livraison_flag else ''
                ]
                
                for col, value in enumerate(row_data, 1):
                    sheet.cell(row=row_num, column=col, value=value)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Save workbook
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"Exported {len(tasks)} tasks to Excel file: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            return False
    
    @staticmethod
    def create_template_excel(file_path: str) -> bool:
        """Create an Excel template file for data entry"""
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Action Plan Template"
            
            # Headers
            headers = [
                'Date', 'PO', 'Catégorie', 'Action', 'Colonne1', 'Customer',
                'Requester', 'Techmac Resp', 'Dead line', 'Status', 'Note',
                'Installation/F', 'Réparation', 'Développement', 'Livraison'
            ]
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add sample data
            sample_data = [
                ['01/01/2024', '202400001', 'Installation', 'Setup new equipment', '', 'Client ABC', 'Manager', 'Amine', '15/01/2024', 'En Cours', 'Priority task', 'X', '', '', ''],
                ['02/01/2024', '202400002', 'Réparation', 'Fix TDR701 machine', '', 'Client XYZ', 'Technician', 'Hassan', '10/01/2024', 'Terminé', 'Completed on time', '', 'X', '', '']
            ]
            
            for row_num, row_data in enumerate(sample_data, 2):
                for col, value in enumerate(row_data, 1):
                    sheet.cell(row=row_num, column=col, value=value)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet
            instructions_sheet = workbook.create_sheet("Instructions")
            instructions = [
                "Instructions pour l'utilisation du template:",
                "",
                "1. Date: Format DD/MM/YYYY (ex: 01/01/2024)",
                "2. PO: Numéro de commande (optionnel)",
                "3. Catégorie: Installation, Réparation, Développement, Livraison, Commercial",
                "4. Action: Description détaillée de la tâche (obligatoire)",
                "5. Customer: Nom du client (obligatoire)",
                "6. Requester: Nom du demandeur (obligatoire)",
                "7. Techmac Resp: Nom du responsable (obligatoire)",
                "8. Dead line: Date d'échéance (format DD/MM/YYYY)",
                "9. Status: En Attente, En Cours, Terminé, Annulé, En Pause",
                "10. Note: Commentaires additionnels",
                "11. Flags: Cocher avec 'X' pour les catégories applicables",
                "",
                "Note: Les colonnes Action, Customer, Requester et Techmac Resp sont obligatoires."
            ]
            
            for row, instruction in enumerate(instructions, 1):
                instructions_sheet.cell(row=row, column=1, value=instruction)
            
            # Save template
            workbook.save(file_path)
            workbook.close()
            
            logger.info(f"Created Excel template: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating Excel template: {str(e)}")
            return False