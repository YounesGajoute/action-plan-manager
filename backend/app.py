#!/usr/bin/env python3
# ===================================================================
# backend/app.py - Main Flask Application
# ===================================================================

import os
import logging
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_migrate import Migrate
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import uuid
from functools import wraps
import requests
from config import Config

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(Config)

# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db)
jwt = JWTManager(app)
CORS(app, origins=["http://localhost:3000", "https://yourdomain.com"])

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s %(message)s',
    handlers=[
        logging.FileHandler('logs/api/app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ===================================================================
# DATABASE MODELS
# ===================================================================

class User(db.Model):
    __tablename__ = 'users'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    email = db.Column(db.String(255), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    microsoft_id = db.Column(db.String(255), unique=True, nullable=True)
    tenant_id = db.Column(db.String(255), nullable=True)
    roles = db.Column(db.JSON, default=lambda: ['user'])
    department = db.Column(db.String(100), nullable=True)
    position = db.Column(db.String(100), nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    last_login = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'email': self.email,
            'name': self.name,
            'roles': self.roles or ['user'],
            'department': self.department,
            'position': self.position,
            'phone': self.phone,
            'is_active': self.is_active,
            'last_login': self.last_login.isoformat() if self.last_login else None,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat()
        }

class Task(db.Model):
    __tablename__ = 'tasks'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    po_number = db.Column(db.String(100), nullable=True)
    date_created = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    category = db.Column(db.String(50), nullable=True)
    action_description = db.Column(db.Text, nullable=False)
    colonne1 = db.Column(db.String(255), nullable=True)  # Additional column from Excel
    customer = db.Column(db.String(255), nullable=False)
    requester = db.Column(db.String(255), nullable=False)
    responsible = db.Column(db.String(255), nullable=False)
    deadline = db.Column(db.DateTime, nullable=True)
    status = db.Column(db.String(50), nullable=False, default='En Attente')
    priority = db.Column(db.String(20), nullable=False, default='Moyen')
    notes = db.Column(db.Text, nullable=True)
    
    # Category flags from Excel
    installation_flag = db.Column(db.Boolean, default=False)
    reparation_flag = db.Column(db.Boolean, default=False)
    developpement_flag = db.Column(db.Boolean, default=False)
    livraison_flag = db.Column(db.Boolean, default=False)
    
    # System fields
    created_by = db.Column(db.String(36), db.ForeignKey('users.id'), nullable=True)
    updated_by = db.Column(db.String(36), db.ForeignKey('users.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'po_number': self.po_number,
            'date_created': self.date_created.isoformat() if self.date_created else None,
            'category': self.category,
            'action_description': self.action_description,
            'colonne1': self.colonne1,
            'customer': self.customer,
            'requester': self.requester,
            'responsible': self.responsible,
            'deadline': self.deadline.isoformat() if self.deadline else None,
            'status': self.status,
            'priority': self.priority,
            'notes': self.notes,
            'installation_flag': self.installation_flag,
            'reparation_flag': self.reparation_flag,
            'developpement_flag': self.developpement_flag,
            'livraison_flag': self.livraison_flag,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat()
        }

class TaskHistory(db.Model):
    __tablename__ = 'task_history'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    task_id = db.Column(db.String(36), db.ForeignKey('tasks.id'), nullable=False)
    field_name = db.Column(db.String(50), nullable=False)
    old_value = db.Column(db.Text, nullable=True)
    new_value = db.Column(db.Text, nullable=True)
    changed_by = db.Column(db.String(36), db.ForeignKey('users.id'), nullable=False)
    changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'task_id': self.task_id,
            'field_name': self.field_name,
            'old_value': self.old_value,
            'new_value': self.new_value,
            'changed_by': self.changed_by,
            'changed_at': self.changed_at.isoformat()
        }

class SyncStatus(db.Model):
    __tablename__ = 'sync_status'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    sync_type = db.Column(db.String(50), nullable=False)  # 'onedrive', 'manual'
    status = db.Column(db.String(20), nullable=False)  # 'success', 'error', 'in_progress'
    message = db.Column(db.Text, nullable=True)
    items_processed = db.Column(db.Integer, default=0)
    items_imported = db.Column(db.Integer, default=0)
    items_updated = db.Column(db.Integer, default=0)
    items_failed = db.Column(db.Integer, default=0)
    started_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime, nullable=True)
    
    def to_dict(self):
        return {
            'id': self.id,
            'sync_type': self.sync_type,
            'status': self.status,
            'message': self.message,
            'items_processed': self.items_processed,
            'items_imported': self.items_imported,
            'items_updated': self.items_updated,
            'items_failed': self.items_failed,
            'started_at': self.started_at.isoformat(),
            'completed_at': self.completed_at.isoformat() if self.completed_at else None
        }

class Notification(db.Model):
    __tablename__ = 'notifications'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    user_id = db.Column(db.String(36), db.ForeignKey('users.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    message = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20), nullable=False, default='info')  # info, success, warning, error
    read = db.Column(db.Boolean, default=False)
    task_id = db.Column(db.String(36), db.ForeignKey('tasks.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'title': self.title,
            'message': self.message,
            'type': self.type,
            'read': self.read,
            'task_id': self.task_id,
            'created_at': self.created_at.isoformat()
        }

# ===================================================================
# UTILITY FUNCTIONS
# ===================================================================

def ms_auth_required(f):
    """Decorator for Microsoft authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'No authorization token provided'}), 401
            
        try:
            # Remove 'Bearer ' prefix
            if token.startswith('Bearer '):
                token = token[7:]
            
            # In a real implementation, validate the Microsoft token
            # For now, we'll create a simple user lookup
            user = User.query.filter_by(microsoft_id=token).first()
            if not user:
                # Create a mock user for development
                user = User(
                    email='user@techmac.ma',
                    name='Test User',
                    microsoft_id=token,
                    roles=['user']
                )
                db.session.add(user)
                db.session.commit()
            
            request.current_user = user
            return f(*args, **kwargs)
            
        except Exception as e:
            logger.error(f"Authentication error: {str(e)}")
            return jsonify({'error': 'Invalid token'}), 401
    
    return decorated_function

def parse_excel_date(date_value):
    """Parse Excel date formats"""
    if not date_value:
        return None
        
    try:
        if isinstance(date_value, str):
            # Handle DD/MM/YY or DD/MM/YYYY format
            if '/' in date_value:
                parts = date_value.split('/')
                if len(parts) == 3:
                    day, month, year = parts
                    # Convert 2-digit year to 4-digit
                    if len(year) == 2:
                        year = '20' + year if int(year) <= 30 else '19' + year
                    return datetime.strptime(f"{day}/{month}/{year}", "%d/%m/%Y")
            else:
                return datetime.strptime(date_value, "%Y-%m-%d")
        elif isinstance(date_value, datetime):
            return date_value
        else:
            # Excel serial number
            from datetime import datetime, timedelta
            return datetime(1900, 1, 1) + timedelta(days=date_value - 2)
    except:
        return None

def normalize_status(status):
    """Normalize status values from Excel"""
    if not status:
        return 'En Attente'
    
    status_map = {
        'done': 'Terminé',
        'completed': 'Terminé',
        'finished': 'Terminé',
        'pending': 'En Attente',
        'waiting': 'En Attente',
        'in-progress': 'En Cours',
        'in progress': 'En Cours',
        'cancelled': 'Annulé',
        'canceled': 'Annulé',
        'on-hold': 'En Pause',
        'on hold': 'En Pause'
    }
    
    return status_map.get(status.lower(), status)

def infer_category_from_flags(row):
    """Infer category from Excel flag columns"""
    if row.get('Installation/F'):
        return 'Installation'
    elif row.get('Réparation'):
        return 'Réparation'
    elif row.get('Développement'):
        return 'Développement'
    elif row.get('Livraison '):
        return 'Livraison'
    elif 'commercial' in str(row.get('Action ', '')).lower():
        return 'Commercial'
    return None

# ===================================================================
# API ROUTES - AUTHENTICATION
# ===================================================================

@app.route('/auth/me', methods=['GET'])
@ms_auth_required
def get_current_user():
    """Get current user information"""
    try:
        user = request.current_user
        return jsonify({
            'success': True,
            'user': user.to_dict()
        })
    except Exception as e:
        logger.error(f"Error getting current user: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/auth/login', methods=['POST'])
def login():
    """Fallback login for development"""
    data = request.get_json()
    email = data.get('email')
    
    if not email:
        return jsonify({'error': 'Email is required'}), 400
    
    user = User.query.filter_by(email=email).first()
    if not user:
        # Create user for development
        user = User(
            email=email,
            name=data.get('name', email.split('@')[0]),
            roles=['user']
        )
        db.session.add(user)
        db.session.commit()
    
    # Update last login
    user.last_login = datetime.utcnow()
    db.session.commit()
    
    # Create JWT token
    token = create_access_token(identity=user.id, expires_delta=timedelta(hours=1))
    
    return jsonify({
        'success': True,
        'user': user.to_dict(),
        'token': token
    })

# ===================================================================
# API ROUTES - TASKS
# ===================================================================

@app.route('/api/tasks', methods=['GET'])
@ms_auth_required
def get_tasks():
    """Get all tasks with optional filtering"""
    try:
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 10, type=int)
        status = request.args.get('status')
        category = request.args.get('category')
        responsible = request.args.get('responsible')
        customer = request.args.get('customer')
        search = request.args.get('search')
        
        # Build query
        query = Task.query
        
        # Apply filters
        if status:
            query = query.filter(Task.status == status)
        if category:
            query = query.filter(Task.category == category)
        if responsible:
            query = query.filter(Task.responsible.ilike(f'%{responsible}%'))
        if customer:
            query = query.filter(Task.customer.ilike(f'%{customer}%'))
        if search:
            query = query.filter(
                db.or_(
                    Task.action_description.ilike(f'%{search}%'),
                    Task.customer.ilike(f'%{search}%'),
                    Task.po_number.ilike(f'%{search}%'),
                    Task.responsible.ilike(f'%{search}%')
                )
            )
        
        # Get total count
        total = query.count()
        
        # Apply pagination
        tasks = query.order_by(Task.created_at.desc()).paginate(
            page=page, per_page=limit, error_out=False
        )
        
        # Calculate counts
        counts = {
            'total': Task.query.count(),
            'pending': Task.query.filter_by(status='En Attente').count(),
            'inProgress': Task.query.filter_by(status='En Cours').count(),
            'completed': Task.query.filter_by(status='Terminé').count(),
            'overdue': Task.query.filter(
                Task.deadline < datetime.utcnow(),
                Task.status.notin_(['Terminé', 'Annulé'])
            ).count()
        }
        
        return jsonify({
            'success': True,
            'tasks': [task.to_dict() for task in tasks.items],
            'counts': counts,
            'total': total,
            'page': page,
            'pages': tasks.pages
        })
        
    except Exception as e:
        logger.error(f"Error getting tasks: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/tasks', methods=['POST'])
@ms_auth_required
def create_task():
    """Create a new task"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['action_description', 'customer', 'requester', 'responsible']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Create task
        task = Task(
            po_number=data.get('po_number'),
            category=data.get('category'),
            action_description=data['action_description'],
            customer=data['customer'],
            requester=data['requester'],
            responsible=data['responsible'],
            deadline=parse_excel_date(data.get('deadline')),
            status=data.get('status', 'En Attente'),
            priority=data.get('priority', 'Moyen'),
            notes=data.get('notes'),
            created_by=request.current_user.id
        )
        
        db.session.add(task)
        db.session.commit()
        
        logger.info(f"Task created: {task.id} by user {request.current_user.id}")
        
        return jsonify({
            'success': True,
            'task': task.to_dict()
        }), 201
        
    except Exception as e:
        logger.error(f"Error creating task: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/tasks/<task_id>', methods=['GET'])
@ms_auth_required
def get_task(task_id):
    """Get a specific task"""
    try:
        task = Task.query.get_or_404(task_id)
        return jsonify({
            'success': True,
            'task': task.to_dict()
        })
    except Exception as e:
        logger.error(f"Error getting task: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/tasks/<task_id>', methods=['PUT'])
@ms_auth_required
def update_task(task_id):
    """Update a task"""
    try:
        task = Task.query.get_or_404(task_id)
        data = request.get_json()
        
        # Track changes for history
        changes = []
        
        # Update fields
        updatable_fields = [
            'po_number', 'category', 'action_description', 'customer',
            'requester', 'responsible', 'status', 'priority', 'notes'
        ]
        
        for field in updatable_fields:
            if field in data:
                old_value = getattr(task, field)
                new_value = data[field]
                
                if field == 'deadline' and new_value:
                    new_value = parse_excel_date(new_value)
                
                if old_value != new_value:
                    changes.append({
                        'field': field,
                        'old_value': str(old_value) if old_value else None,
                        'new_value': str(new_value) if new_value else None
                    })
                    setattr(task, field, new_value)
        
        # Handle deadline separately
        if 'deadline' in data:
            old_deadline = task.deadline
            new_deadline = parse_excel_date(data['deadline'])
            if old_deadline != new_deadline:
                changes.append({
                    'field': 'deadline',
                    'old_value': old_deadline.isoformat() if old_deadline else None,
                    'new_value': new_deadline.isoformat() if new_deadline else None
                })
                task.deadline = new_deadline
        
        task.updated_by = request.current_user.id
        task.updated_at = datetime.utcnow()
        
        # Save changes
        db.session.commit()
        
        # Record history
        for change in changes:
            history = TaskHistory(
                task_id=task.id,
                field_name=change['field'],
                old_value=change['old_value'],
                new_value=change['new_value'],
                changed_by=request.current_user.id
            )
            db.session.add(history)
        
        db.session.commit()
        
        logger.info(f"Task updated: {task.id} by user {request.current_user.id}")
        
        return jsonify({
            'success': True,
            'task': task.to_dict()
        })
        
    except Exception as e:
        logger.error(f"Error updating task: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/tasks/<task_id>', methods=['DELETE'])
@ms_auth_required
def delete_task(task_id):
    """Delete a task"""
    try:
        task = Task.query.get_or_404(task_id)
        
        # Delete related history
        TaskHistory.query.filter_by(task_id=task_id).delete()
        
        # Delete task
        db.session.delete(task)
        db.session.commit()
        
        logger.info(f"Task deleted: {task_id} by user {request.current_user.id}")
        
        return jsonify({
            'success': True,
            'message': 'Task deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting task: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Internal server error'}), 500

# ===================================================================
# API ROUTES - EXCEL IMPORT
# ===================================================================

@app.route('/api/tasks/import', methods=['POST'])
@ms_auth_required
def import_tasks():
    """Import tasks from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file'}), 400
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Process Excel file
        imported_count = 0
        updated_count = 0
        errors = []
        
        try:
            # Read Excel file
            workbook = load_workbook(filepath)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Process each row
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Create row dictionary
                    row_data = dict(zip(headers, row))
                    
                    # Skip if no Action description
                    if not row_data.get('Action '):
                        continue
                    
                    # Parse data
                    date_created = parse_excel_date(row_data.get('Date'))
                    deadline = parse_excel_date(row_data.get('Dead line '))
                    
                    # Check if task exists (by PO number and action)
                    existing_task = None
                    if row_data.get('PO'):
                        existing_task = Task.query.filter_by(
                            po_number=row_data['PO'],
                            action_description=row_data['Action ']
                        ).first()
                    
                    if existing_task:
                        # Update existing task
                        existing_task.customer = row_data.get('Customer', '')
                        existing_task.requester = row_data.get('Requester', '')
                        existing_task.responsible = row_data.get('Techmac Resp', '')
                        existing_task.status = normalize_status(row_data.get('Status'))
                        existing_task.notes = row_data.get('Note')
                        existing_task.deadline = deadline
                        existing_task.updated_at = datetime.utcnow()
                        
                        updated_count += 1
                    else:
                        # Create new task
                        task = Task(
                            po_number=row_data.get('PO'),
                            date_created=date_created or datetime.utcnow(),
                            category=row_data.get('Catégorie') or infer_category_from_flags(row_data),
                            action_description=row_data.get('Action ', ''),
                            colonne1=row_data.get('Colonne1'),
                            customer=row_data.get('Customer', ''),
                            requester=row_data.get('Requester', ''),
                            responsible=row_data.get('Techmac Resp', ''),
                            deadline=deadline,
                            status=normalize_status(row_data.get('Status')),
                            notes=row_data.get('Note'),
                            installation_flag=bool(row_data.get('Installation/F')),
                            reparation_flag=bool(row_data.get('Réparation')),
                            developpement_flag=bool(row_data.get('Développement')),
                            livraison_flag=bool(row_data.get('Livraison ')),
                            created_by=request.current_user.id
                        )
                        
                        db.session.add(task)
                        imported_count += 1
                
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
            
            # Commit all changes
            db.session.commit()
            
            # Record sync status
            sync_status = SyncStatus(
                sync_type='manual',
                status='success',
                message=f'Imported {imported_count} tasks, updated {updated_count} tasks',
                items_processed=imported_count + updated_count,
                items_imported=imported_count,
                items_updated=updated_count,
                items_failed=len(errors),
                completed_at=datetime.utcnow()
            )
            db.session.add(sync_status)
            db.session.commit()
            
        finally:
            # Clean up temporary file
            if os.path.exists(filepath):
                os.remove(filepath)
        
        logger.info(f"Excel import completed: {imported_count} imported, {updated_count} updated, {len(errors)} errors")
        
        return jsonify({
            'success': True,
            'imported': imported_count,
            'updated': updated_count,
            'errors': errors
        })
        
    except Exception as e:
        logger.error(f"Error importing Excel: {str(e)}")
        return jsonify({'error': 'Import failed'}), 500

# ===================================================================
# API ROUTES - ANALYTICS
# ===================================================================

@app.route('/api/analytics/dashboard', methods=['GET'])
@ms_auth_required
def get_dashboard_analytics():
    """Get dashboard analytics data"""
    try:
        # Task counts
        total_tasks = Task.query.count()
        completed_tasks = Task.query.filter_by(status='Terminé').count()
        in_progress_tasks = Task.query.filter_by(status='En Cours').count()
        pending_tasks = Task.query.filter_by(status='En Attente').count()
        overdue_tasks = Task.query.filter(
            Task.deadline < datetime.utcnow(),
            Task.status.notin_(['Terminé', 'Annulé'])
        ).count()
        
        # Category distribution
        categories = db.session.query(
            Task.category,
            db.func.count(Task.id).label('count')
        ).group_by(Task.category).all()
        
        category_distribution = []
        for category, count in categories:
            if category:
                percentage = round((count / total_tasks) * 100, 1) if total_tasks > 0 else 0
                category_distribution.append({
                    'category': category,
                    'count': count,
                    'percentage': percentage
                })
        
        # Weekly progress (last 7 days)
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        weekly_progress = []
        
        for i in range(7):
            date = seven_days_ago + timedelta(days=i)
            date_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
            date_end = date_start + timedelta(days=1)
            
            created = Task.query.filter(
                Task.created_at >= date_start,
                Task.created_at < date_end
            ).count()
            
            completed = Task.query.filter(
                Task.updated_at >= date_start,
                Task.updated_at < date_end,
                Task.status == 'Terminé'
            ).count()
            
            weekly_progress.append({
                'date': date.strftime('%a'),
                'created': created,
                'completed': completed
            })
        
        # Team performance
        team_performance = db.session.query(
            Task.responsible,
            db.func.count(Task.id).label('total'),
            db.func.sum(db.case([(Task.status == 'Terminé', 1)], else_=0)).label('completed'),
            db.func.sum(db.case([(Task.status.in_(['En Attente', 'En Cours']), 1)], else_=0)).label('pending')
        ).group_by(Task.responsible).all()
        
        team_data = []
        for responsible, total, completed, pending in team_performance:
            completion_rate = round((completed / total) * 100, 1) if total > 0 else 0
            team_data.append({
                'responsible': responsible,
                'total': total,
                'completed': completed or 0,
                'pending': pending or 0,
                'completionRate': completion_rate
            })
        
        # Calculate changes (mock data for demo)
        changes = {
            'totalTasks': 5,  # +5% vs last week
            'completedTasks': 12,  # +12% vs last week
            'inProgressTasks': -3,  # -3% vs last week
            'overdueTasks': -8  # -8% vs last week
        }
        
        return jsonify({
            'success': True,
            'taskCounts': {
                'total': total_tasks,
                'completed': completed_tasks,
                'inProgress': in_progress_tasks,
                'pending': pending_tasks,
                'overdue': overdue_tasks
            },
            'categoryDistribution': category_distribution,
            'weeklyProgress': weekly_progress,
            'teamPerformance': team_data,
            'changes': changes
        })
        
    except Exception as e:
        logger.error(f"Error getting dashboard analytics: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

# ===================================================================
# API ROUTES - NOTIFICATIONS
# ===================================================================

@app.route('/api/notifications', methods=['GET'])
@ms_auth_required
def get_notifications():
    """Get user notifications"""
    try:
        notifications = Notification.query.filter_by(
            user_id=request.current_user.id
        ).order_by(Notification.created_at.desc()).limit(50).all()
        
        return jsonify({
            'success': True,
            'notifications': [notif.to_dict() for notif in notifications]
        })
        
    except Exception as e:
        logger.error(f"Error getting notifications: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/notifications/<notification_id>/read', methods=['PUT'])
@ms_auth_required
def mark_notification_read(notification_id):
    """Mark notification as read"""
    try:
        notification = Notification.query.filter_by(
            id=notification_id,
            user_id=request.current_user.id
        ).first_or_404()
        
        notification.read = True
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        logger.error(f"Error marking notification as read: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/notifications/read-all', methods=['PUT'])
@ms_auth_required
def mark_all_notifications_read():
    """Mark all notifications as read"""
    try:
        Notification.query.filter_by(
            user_id=request.current_user.id,
            read=False
        ).update({'read': True})
        
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        logger.error(f"Error marking all notifications as read: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

# ===================================================================
# API ROUTES - SYNC STATUS
# ===================================================================

@app.route('/api/sync/status', methods=['GET'])
@ms_auth_required
def get_sync_status():
    """Get synchronization status"""
    try:
        # Get latest sync status
        latest_sync = SyncStatus.query.order_by(SyncStatus.started_at.desc()).first()
        
        # Get sync history (last 10)
        sync_history = SyncStatus.query.order_by(
            SyncStatus.started_at.desc()
        ).limit(10).all()
        
        # Mock OneDrive status for demo
        is_online = True
        last_sync_time = latest_sync.completed_at if latest_sync else None
        
        return jsonify({
            'success': True,
            'isOnlineSyncActive': is_online,
            'lastSyncTime': last_sync_time.isoformat() if last_sync_time else None,
            'syncInProgress': latest_sync.status == 'in_progress' if latest_sync else False,
            'error': None,
            'syncHistory': [sync.to_dict() for sync in sync_history]
        })
        
    except Exception as e:
        logger.error(f"Error getting sync status: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/sync/trigger', methods=['POST'])
@ms_auth_required
def trigger_sync():
    """Trigger manual synchronization"""
    try:
        # Create sync status record
        sync_status = SyncStatus(
            sync_type='manual',
            status='in_progress',
            message='Manual sync triggered'
        )
        db.session.add(sync_status)
        db.session.commit()
        
        # In a real implementation, this would trigger the actual sync process
        # For demo, we'll just mark it as completed
        sync_status.status = 'success'
        sync_status.message = 'Manual sync completed successfully'
        sync_status.completed_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        logger.error(f"Error triggering sync: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

# ===================================================================
# API ROUTES - HEALTH CHECK
# ===================================================================

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Check database connection
        db.session.execute('SELECT 1')
        
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.utcnow().isoformat(),
            'version': '1.0.0',
            'services': {
                'database': 'healthy',
                'api': 'healthy'
            }
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat()
        }), 500

@app.route('/health/db', methods=['GET'])
def health_check_db():
    """Database health check"""
    try:
        db.session.execute('SELECT 1')
        return jsonify({'status': 'healthy', 'database': 'connected'})
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 500

# ===================================================================
# BACKGROUND TASKS (for future implementation)
# ===================================================================

def create_notification(user_id, title, message, notification_type='info', task_id=None):
    """Create a notification for a user"""
    try:
        notification = Notification(
            user_id=user_id,
            title=title,
            message=message,
            type=notification_type,
            task_id=task_id
        )
        db.session.add(notification)
        db.session.commit()
        return notification
    except Exception as e:
        logger.error(f"Error creating notification: {str(e)}")
        return None

def check_overdue_tasks():
    """Check for overdue tasks and create notifications"""
    try:
        overdue_tasks = Task.query.filter(
            Task.deadline < datetime.utcnow(),
            Task.status.notin_(['Terminé', 'Annulé'])
        ).all()
        
        for task in overdue_tasks:
            # Find users to notify (responsible person, managers, etc.)
            users_to_notify = User.query.filter(
                db.or_(
                    User.name == task.responsible,
                    User.roles.contains(['admin']),
                    User.roles.contains(['manager'])
                )
            ).all()
            
            for user in users_to_notify:
                create_notification(
                    user_id=user.id,
                    title='Tâche en retard',
                    message=f'La tâche "{task.action_description}" est en retard (échéance: {task.deadline.strftime("%d/%m/%Y")})',
                    notification_type='warning',
                    task_id=task.id
                )
        
        logger.info(f"Checked overdue tasks: {len(overdue_tasks)} found")
        
    except Exception as e:
        logger.error(f"Error checking overdue tasks: {str(e)}")

# ===================================================================
# ERROR HANDLERS
# ===================================================================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Resource not found'}), 404

@app.errorhandler(400)
def bad_request(error):
    return jsonify({'error': 'Bad request'}), 400

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    logger.error(f"Internal server error: {str(error)}")
    return jsonify({'error': 'Internal server error'}), 500

# ===================================================================
# APPLICATION INITIALIZATION
# ===================================================================

def create_tables():
    """Create database tables"""
    with app.app_context():
        try:
            db.create_all()
            logger.info("Database tables created successfully")
            
            # Create default admin user if it doesn't exist
            admin_user = User.query.filter_by(email='admin@techmac.ma').first()
            if not admin_user:
                admin_user = User(
                    email='admin@techmac.ma',
                    name='Administrator',
                    roles=['admin', 'manager', 'user'],
                    department='IT',
                    position='System Administrator'
                )
                db.session.add(admin_user)
                db.session.commit()
                logger.info("Default admin user created")
                
        except Exception as e:
            logger.error(f"Error creating tables: {str(e)}")

# ===================================================================
# MAIN APPLICATION ENTRY POINT
# ===================================================================

if __name__ == '__main__':
    # Ensure upload directory exists
    os.makedirs(app.config.get('UPLOAD_FOLDER', 'uploads'), exist_ok=True)
    os.makedirs('logs/api', exist_ok=True)
    
    # Create database tables
    create_tables()
    
    # Run the application
    debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(
        host='0.0.0.0',
        port=int(os.getenv('API_PORT', 5000)),
        debug=debug_mode
    )