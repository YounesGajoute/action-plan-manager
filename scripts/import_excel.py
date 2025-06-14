#!/usr/bin/env python3
# ===================================================================
# scripts/import_excel.py - Enhanced Excel Import Script
# ===================================================================

import os
import sys
import logging
import argparse
import traceback
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sqlalchemy
from sqlalchemy.exc import IntegrityError
import re

# Add the parent directory to the path to import app modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app, db
from app.models import Task, User, ImportLog, SyncStatus
from app.services.excel_service import ExcelService
from app.utils.validators import validate_email, validate_date, validate_po_number

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/import_excel.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class ExcelImporter:
    """Enhanced Excel importer with validation, progress tracking, and error handling"""
    
    # Required columns mapping
    COLUMN_MAPPING = {
        'Date': 'date',
        'PO': 'po_number', 
        'Catégorie': 'category',
        'Action': 'action_description',
        'Customer': 'customer',
        'Requester': 'requester',
        'Techmac Resp': 'responsible',
        'Dead line': 'deadline',
        'Status': 'status',
        'Note': 'notes'
    }
    
    # Required columns
    REQUIRED_COLUMNS = ['Date', 'Action', 'Customer', 'Requester', 'Techmac Resp']
    
    # Valid categories
    VALID_CATEGORIES = [
        'Installation', 'Réparation', 'Développement', 'Livraison', 'Commercial'
    ]
    
    # Valid statuses
    VALID_STATUSES = [
        'En Attente', 'En Cours', 'Terminé', 'Annulé', 'En Pause'
    ]
    
    def __init__(self, file_path: str, dry_run: bool = False, batch_size: int = 100):
        self.file_path = file_path
        self.dry_run = dry_run
        self.batch_size = batch_size
        self.errors = []
        self.warnings = []
        self.stats = {
            'total_rows': 0,
            'processed': 0,
            'imported': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0
        }
        
        # Initialize Flask app context
        self.app = create_app()
        self.app_context = self.app.app_context()
        
    def __enter__(self):
        self.app_context.push()
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.app_context.pop()
        
    def validate_file(self) -> Tuple[bool, str]:
        """Validate Excel file structure and accessibility"""
        try:
            if not os.path.exists(self.file_path):
                return False, f"File not found: {self.file_path}"
                
            if not self.file_path.lower().endswith(('.xlsx', '.xls')):
                return False, f"Invalid file format. Expected .xlsx or .xls, got: {self.file_path}"
                
            # Try to open the file
            workbook = load_workbook(self.file_path, read_only=True)
            sheet = workbook.active
            
            if sheet.max_row < 2:
                return False, "File appears to be empty (no data rows)"
                
            # Get headers
            headers = [cell.value for cell in sheet[1] if cell.value]
            
            # Check for required columns
            missing_columns = []
            for required_col in self.REQUIRED_COLUMNS:
                if required_col not in headers:
                    missing_columns.append(required_col)
                    
            if missing_columns:
                return False, f"Missing required columns: {', '.join(missing_columns)}"
                
            # Log file info
            logger.info(f"File validation passed:")
            logger.info(f"  - Headers: {headers}")
            logger.info(f"  - Data rows: {sheet.max_row - 1}")
            
            workbook.close()
            return True, "File validation successful"
            
        except Exception as e:
            return False, f"File validation error: {str(e)}"
            
    def read_excel_file(self) -> Optional[pd.DataFrame]:
        """Read Excel file with proper error handling"""
        try:
            logger.info(f"Reading Excel file: {self.file_path}")
            
            # Read with pandas for better data handling
            df = pd.read_excel(
                self.file_path,
                engine='openpyxl',
                sheet_name=0,
                header=0,
                dtype=str,  # Read everything as string initially
                na_values=['', 'N/A', 'NA', 'null', 'NULL']
            )
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            # Strip whitespace from string columns
            for col in df.select_dtypes(include=['object']).columns:
                df[col] = df[col].astype(str).str.strip()
                
            # Replace 'nan' strings with None
            df = df.replace('nan', None)
            
            self.stats['total_rows'] = len(df)
            logger.info(f"Successfully read {len(df)} rows from Excel file")
            
            return df
            
        except Exception as e:
            error_msg = f"Error reading Excel file: {str(e)}"
            logger.error(error_msg)
            self.errors.append(error_msg)
            return None
            
    def validate_row(self, row: pd.Series, row_num: int) -> Tuple[bool, Dict[str, Any], List[str]]:
        """Validate a single row of data"""
        errors = []
        validated_data = {}
        
        try:
            # Validate date
            date_value = row.get('Date')
            if pd.isna(date_value) or not date_value:
                errors.append(f"Row {row_num}: Date is required")
            else:
                try:
                    if isinstance(date_value, str):
                        # Try different date formats
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
                            try:
                                validated_data['date'] = datetime.strptime(date_value, fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            errors.append(f"Row {row_num}: Invalid date format: {date_value}")
                    else:
                        validated_data['date'] = pd.to_datetime(date_value).date()
                except:
                    errors.append(f"Row {row_num}: Invalid date: {date_value}")
                    
            # Validate PO number
            po_number = row.get('PO')
            if po_number and not pd.isna(po_number):
                if not validate_po_number(str(po_number)):
                    errors.append(f"Row {row_num}: Invalid PO number format: {po_number}")
                else:
                    validated_data['po_number'] = str(po_number).upper()
                    
            # Validate required text fields
            for field in ['Action', 'Customer', 'Requester', 'Techmac Resp']:
                value = row.get(field)
                if pd.isna(value) or not value or str(value).strip() == '':
                    errors.append(f"Row {row_num}: {field} is required")
                else:
                    db_field = self.COLUMN_MAPPING.get(field, field.lower())
                    validated_data[db_field] = str(value).strip()
                    
            # Validate category
            category = row.get('Catégorie')
            if category and not pd.isna(category):
                if category not in self.VALID_CATEGORIES:
                    self.warnings.append(f"Row {row_num}: Unknown category '{category}', will use as-is")
                validated_data['category'] = str(category)
                
            # Validate status
            status = row.get('Status')
            if status and not pd.isna(status):
                if status not in self.VALID_STATUSES:
                    self.warnings.append(f"Row {row_num}: Unknown status '{status}', will use as-is")
                validated_data['status'] = str(status)
            else:
                validated_data['status'] = 'En Attente'  # Default status
                
            # Validate deadline
            deadline = row.get('Dead line')
            if deadline and not pd.isna(deadline):
                try:
                    if isinstance(deadline, str):
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                validated_data['deadline'] = datetime.strptime(deadline, fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            errors.append(f"Row {row_num}: Invalid deadline format: {deadline}")
                    else:
                        validated_data['deadline'] = pd.to_datetime(deadline).date()
                except:
                    errors.append(f"Row {row_num}: Invalid deadline: {deadline}")
                    
            # Add notes
            notes = row.get('Note')
            if notes and not pd.isna(notes):
                validated_data['notes'] = str(notes)
                
            # Add metadata
            validated_data['created_at'] = datetime.utcnow()
            validated_data['updated_at'] = datetime.utcnow()
            
            return len(errors) == 0, validated_data, errors
            
        except Exception as e:
            errors.append(f"Row {row_num}: Validation error: {str(e)}")
            return False, {}, errors
            
    def check_duplicate(self, validated_data: Dict[str, Any]) -> Optional[Task]:
        """Check if task already exists based on PO number or unique combination"""
        try:
            # First check by PO number if provided
            if validated_data.get('po_number'):
                existing = Task.query.filter_by(po_number=validated_data['po_number']).first()
                if existing:
                    return existing
                    
            # Check by combination of key fields
            existing = Task.query.filter(
                Task.action_description == validated_data['action_description'],
                Task.customer == validated_data['customer'],
                Task.date == validated_data['date']
            ).first()
            
            return existing
            
        except Exception as e:
            logger.error(f"Error checking for duplicates: {str(e)}")
            return None
            
    def import_row(self, validated_data: Dict[str, Any], row_num: int) -> Tuple[bool, str]:
        """Import a single validated row"""
        try:
            # Check for duplicates
            existing_task = self.check_duplicate(validated_data)
            
            if existing_task:
                if not self.dry_run:
                    # Update existing task
                    for key, value in validated_data.items():
                        if key not in ['created_at']:  # Don't update created_at
                            setattr(existing_task, key, value)
                    existing_task.updated_at = datetime.utcnow()
                    
                self.stats['updated'] += 1
                return True, f"Updated existing task (ID: {existing_task.id if existing_task else 'N/A'})"
            else:
                if not self.dry_run:
                    # Create new task
                    new_task = Task(**validated_data)
                    db.session.add(new_task)
                    
                self.stats['imported'] += 1
                return True, "Created new task"
                
        except Exception as e:
            error_msg = f"Import error: {str(e)}"
            logger.error(f"Row {row_num}: {error_msg}")
            return False, error_msg
            
    def process_batch(self, df_batch: pd.DataFrame, start_row: int) -> None:
        """Process a batch of rows"""
        batch_errors = []
        
        for idx, row in df_batch.iterrows():
            row_num = start_row + idx
            self.stats['processed'] += 1
            
            # Validate row
            is_valid, validated_data, row_errors = self.validate_row(row, row_num)
            
            if not is_valid:
                self.stats['errors'] += 1
                self.stats['skipped'] += 1
                batch_errors.extend(row_errors)
                continue
                
            # Import row
            success, message = self.import_row(validated_data, row_num)
            
            if not success:
                self.stats['errors'] += 1
                self.stats['skipped'] += 1
                batch_errors.append(f"Row {row_num}: {message}")
            
            # Progress reporting
            if self.stats['processed'] % 50 == 0:
                progress = (self.stats['processed'] / self.stats['total_rows']) * 100
                logger.info(f"Progress: {progress:.1f}% ({self.stats['processed']}/{self.stats['total_rows']})")
                
        # Add batch errors to main error list
        self.errors.extend(batch_errors)
        
        # Commit batch if not dry run
        if not self.dry_run and not batch_errors:
            try:
                db.session.commit()
                logger.info(f"Committed batch of {len(df_batch)} rows")
            except Exception as e:
                db.session.rollback()
                logger.error(f"Failed to commit batch: {str(e)}")
                self.errors.append(f"Batch commit failed: {str(e)}")
                
    def run_import(self) -> Dict[str, Any]:
        """Run the complete import process"""
        start_time = datetime.utcnow()
        
        try:
            logger.info(f"Starting Excel import: {self.file_path}")
            logger.info(f"Mode: {'DRY RUN' if self.dry_run else 'LIVE IMPORT'}")
            
            # Validate file
            is_valid, message = self.validate_file()
            if not is_valid:
                return {'success': False, 'error': message}
                
            # Read Excel file
            df = self.read_excel_file()
            if df is None:
                return {'success': False, 'error': 'Failed to read Excel file'}
                
            # Process in batches
            total_batches = (len(df) + self.batch_size - 1) // self.batch_size
            logger.info(f"Processing {len(df)} rows in {total_batches} batches of {self.batch_size}")
            
            for batch_num in range(total_batches):
                start_idx = batch_num * self.batch_size
                end_idx = min(start_idx + self.batch_size, len(df))
                
                logger.info(f"Processing batch {batch_num + 1}/{total_batches} (rows {start_idx + 2}-{end_idx + 1})")
                
                batch_df = df.iloc[start_idx:end_idx]
                self.process_batch(batch_df, start_idx + 2)  # +2 for header and 0-indexing
                
            # Final commit
            if not self.dry_run:
                try:
                    db.session.commit()
                    logger.info("Final commit completed")
                except Exception as e:
                    db.session.rollback()
                    logger.error(f"Final commit failed: {str(e)}")
                    return {'success': False, 'error': f"Final commit failed: {str(e)}"}
                    
            # Calculate results
            end_time = datetime.utcnow()
            duration = (end_time - start_time).total_seconds()
            
            result = {
                'success': True,
                'stats': self.stats,
                'errors': self.errors,
                'warnings': self.warnings,
                'duration_seconds': duration,
                'dry_run': self.dry_run
            }
            
            logger.info(f"Import completed in {duration:.2f} seconds")
            logger.info(f"Stats: {self.stats}")
            
            return result
            
        except Exception as e:
            if not self.dry_run:
                db.session.rollback()
            error_msg = f"Import failed: {str(e)}"
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            return {'success': False, 'error': error_msg}

def print_usage():
    """Print usage instructions"""
    print("""
Usage: python scripts/import_excel.py <path_to_excel_file> [options]

Examples:
    python scripts/import_excel.py data/Plan_daction.xlsx
    python scripts/import_excel.py /path/to/file.xlsx --dry-run
    python scripts/import_excel.py data/file.xlsx --batch-size 50

Options:
    --dry-run         Only validate and preview import without making changes
    --batch-size N    Process N rows at a time (default: 100)
    --help           Show this help message

The Excel file should have the following required columns:
- Date: Task creation date (DD/MM/YYYY format)
- Action: Task description (required)
- Customer: Customer name (required)
- Requester: Person who requested the task (required)
- Techmac Resp: Responsible person (required)

Optional columns:
- PO: Purchase Order number
- Catégorie: Task category
- Dead line: Task deadline (DD/MM/YYYY format)
- Status: Task status
- Note: Additional notes
    """)

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description='Import Excel file into Action Plan database')
    parser.add_argument('file_path', help='Path to Excel file')
    parser.add_argument('--dry-run', action='store_true', help='Validate only, do not import')
    parser.add_argument('--batch-size', type=int, default=100, help='Batch size for processing')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
        
    # Validate arguments
    if not os.path.exists(args.file_path):
        print(f"Error: File not found: {args.file_path}")
        sys.exit(1)
        
    try:
        with ExcelImporter(args.file_path, args.dry_run, args.batch_size) as importer:
            result = importer.run_import()
            
            print("\n" + "="*60)
            print("IMPORT RESULTS")
            print("="*60)
            
            if result['success']:
                stats = result['stats']
                print(f"✅ Import {'simulation' if result['dry_run'] else 'completed'} successfully!")
                print(f"📊 Statistics:")
                print(f"   Total rows processed: {stats['processed']}")
                print(f"   New tasks imported: {stats['imported']}")
                print(f"   Existing tasks updated: {stats['updated']}")
                print(f"   Rows skipped (errors): {stats['skipped']}")
                print(f"   Duration: {result['duration_seconds']:.2f} seconds")
                
                if result['warnings']:
                    print(f"\n⚠️  Warnings ({len(result['warnings'])}):")
                    for warning in result['warnings'][:10]:  # Show first 10
                        print(f"   {warning}")
                    if len(result['warnings']) > 10:
                        print(f"   ... and {len(result['warnings']) - 10} more warnings")
                        
                if result['errors']:
                    print(f"\n❌ Errors ({len(result['errors'])}):")
                    for error in result['errors'][:10]:  # Show first 10
                        print(f"   {error}")
                    if len(result['errors']) > 10:
                        print(f"   ... and {len(result['errors']) - 10} more errors")
                        
                if result['dry_run']:
                    print(f"\n💡 This was a dry run. Use without --dry-run to actually import the data.")
                    
            else:
                print(f"❌ Import failed: {result['error']}")
                sys.exit(1)
                
    except KeyboardInterrupt:
        print("\n⚠️  Import interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Unexpected error: {str(e)}")
        logger.error(traceback.format_exc())
        sys.exit(1)

if __name__ == '__main__':
    main()